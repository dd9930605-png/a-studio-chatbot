#!/usr/bin/env python3
"""Generate questionnaire Excel template for thesis experiment."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "questionnaire-template.xlsx"

HEADERS = [
    "sectionOrder",
    "sectionTitle",
    "variableName",
    "itemId",
    "questionText",
    "questionType",
    "scaleMin",
    "scaleMax",
    "scaleLabelMin",
    "scaleLabelMax",
    "required",
    "order",
    "reverseScored",
    "notes",
]

EXAMPLE_ROWS = [
    # sectionOrder, sectionTitle, variableName, itemId, questionText, questionType, ...
    [
        1,
        "AI 推薦帶來的意外感受",
        "surprise",
        "surprise_1",
        "我覺得此 AI 穿搭助理推薦的穿搭是一個好的驚喜。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        1,
        "FALSE",
        "範例題，可依論文量表修改",
    ],
    [
        1,
        "AI 推薦帶來的意外感受",
        "surprise",
        "surprise_2",
        "AI 推薦的穿搭與我原本預期的穿搭不同，但我覺得這樣的推薦是令人愉快的。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        2,
        "FALSE",
        "",
    ],
    [
        2,
        "學習體驗",
        "learning",
        "learning_1",
        "透過與 AI 穿搭助理的互動，我學到了如何為面試選擇合適的穿搭。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        3,
        "FALSE",
        "範例題",
    ],
    [
        2,
        "學習體驗",
        "learning",
        "learning_2",
        "AI 穿搭助理提供的建議讓我更了解面試穿搭的考量重點。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        4,
        "FALSE",
        "",
    ],
    [
        3,
        "認知負荷",
        "cognitive_load",
        "cogload_1",
        "與 AI 穿搭助理互動的過程讓我覺得費力。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        5,
        "TRUE",
        "反向計分範例",
    ],
    [
        3,
        "認知負荷",
        "cognitive_load",
        "cogload_2",
        "我需要花很多心力才能理解 AI 穿搭助理的建議。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        6,
        "TRUE",
        "",
    ],
    [
        4,
        "使用意願",
        "intention",
        "intention_1",
        "未來若有類似需求，我願意再次使用此 AI 穿搭助理。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        7,
        "FALSE",
        "範例題",
    ],
    [
        4,
        "使用意願",
        "intention",
        "intention_2",
        "我會向他人推薦此 AI 穿搭助理。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        8,
        "FALSE",
        "",
    ],
    [
        5,
        "主觀掌控感",
        "control",
        "control_1",
        "在與 AI 穿搭助理互動的過程中，我覺得自己能掌控推薦結果。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        9,
        "FALSE",
        "範例題",
    ],
    [
        5,
        "主觀掌控感",
        "control",
        "control_2",
        "我覺得 AI 穿搭助理有認真考慮我提供的資訊。",
        "likert_7",
        1,
        7,
        "非常不同意",
        "非常同意",
        "TRUE",
        10,
        "FALSE",
        "",
    ],
    [
        6,
        "整體評價",
        "overall",
        "overall_open",
        "請簡述你對此次 AI 穿搭助理體驗的整體感受（選填）。",
        "text",
        "",
        "",
        "",
        "",
        "FALSE",
        11,
        "FALSE",
        "開放題範例，可刪除",
    ],
]

INSTRUCTIONS = [
    ["問卷 Excel 範本 — 填寫說明", ""],
    ["", ""],
    ["用途", "請在此檔「問卷題目」工作表填寫正式問卷。完成後交給開發者轉成系統內建問卷。"],
    ["", ""],
    ["不在此問卷內的項目", "「使用 AI 前最喜歡哪一套穿搭」屬於 AI 前置問題，已在系統 /pre 頁面處理，請勿重複加入。"],
    ["", ""],
    ["欄位說明", ""],
    ["sectionOrder", "構面順序（數字，越小越前面）"],
    ["sectionTitle", "構面標題（顯示給受試者看的區塊名稱）"],
    ["variableName", "變項名稱（英文，供 SPSS / 匯出用，同一構面共用）"],
    ["itemId", "題目唯一 ID（英文底線，如 surprise_1）"],
    ["questionText", "題目文字（顯示給受試者）"],
    ["questionType", "題型：likert_7（七點量表）/ text（開放題）/ single_choice（單選，需另加選項欄）"],
    ["scaleMin / scaleMax", "量表範圍，likert_7 固定填 1 和 7"],
    ["scaleLabelMin / scaleLabelMax", "量表端點文字，如「非常不同意」「非常同意」"],
    ["required", "是否必填：TRUE 或 FALSE"],
    ["order", "題目顯示順序（全卷連號）"],
    ["reverseScored", "是否反向計分：TRUE 或 FALSE（供 SPSS 分析備註）"],
    ["notes", "備註（僅供研究者，不顯示給受試者）"],
    ["", ""],
    ["七點量表", "1 = 非常不同意，2 = 不同意，3 = 稍微不同意，4 = 中立，5 = 稍微同意，6 = 同意，7 = 非常同意"],
    ["", ""],
    ["注意", "「可接受 3 套穿搭」步驟已取消，請勿加入相關題目。"],
    ["", ""],
    ["完成後", "儲存檔案並提供給開發者，將轉為 public/questionnaire.json 並整合至 /survey 頁面。"],
]


def style_header_row(ws, row: int, col_count: int) -> None:
    fill = PatternFill("solid", fgColor="4F46E5")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_width: int = 10, max_width: int = 48) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)) + 2, max_width))
        ws.column_dimensions[letter].width = max_len


def build_workbook() -> Workbook:
    wb = Workbook()

    # Instructions sheet
    ws_info = wb.active
    ws_info.title = "填寫說明"
    for row_idx, (a, b) in enumerate(INSTRUCTIONS, start=1):
        ws_info.cell(row=row_idx, column=1, value=a)
        ws_info.cell(row=row_idx, column=2, value=b)
    ws_info.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_info.column_dimensions["A"].width = 22
    ws_info.column_dimensions["B"].width = 72

    # Questions sheet
    ws_q = wb.create_sheet("問卷題目")
    for col_idx, header in enumerate(HEADERS, start=1):
        ws_q.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_q, 1, len(HEADERS))

    for row_idx, row_data in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_q.cell(row=row_idx, column=col_idx, value=value)

    ws_q.freeze_panes = "A2"
    auto_width(ws_q)

    # Empty template row hint
    hint_row = len(EXAMPLE_ROWS) + 3
    ws_q.cell(row=hint_row, column=1, value="↓ 請在下方新增或修改你的正式題目 ↓")
    ws_q.merge_cells(start_row=hint_row, start_column=1, end_row=hint_row, end_column=len(HEADERS))
    ws_q.cell(row=hint_row, column=1).font = Font(italic=True, color="6B7280")

    return wb


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
